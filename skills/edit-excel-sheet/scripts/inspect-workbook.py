# /// script
# requires-python = ">=3.12"
# dependencies = ["openpyxl>=3.1"]
# ///
"""Dump the structure of an .xlsx/.xlsm workbook so edits can be planned safely.

Reports per sheet: dimensions, headers, merged ranges, formula count, and
features (charts, images, tables) that openpyxl handles imperfectly, so the
agent knows what a load/save cycle could disturb before touching the file.
"""

import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def main() -> int:
    if len(sys.argv) != 2:
        print(f"Usage: uv run {sys.argv[0]} <workbook.xlsx|.xlsm>")
        return 2
    path = Path(sys.argv[1])
    if not path.exists():
        print(f"Error: file not found: {path}")
        return 1
    if path.suffix.lower() not in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        print(f"Error: {path.suffix} is not an OOXML workbook; this tool reads .xlsx/.xlsm.")
        print("Legacy .xls files need a different library (xlrd) or conversion first.")
        return 1

    # keep_vba load would be needed for editing .xlsm; for inspection plain load is fine
    wb = load_workbook(path, data_only=False)
    print(f"Workbook: {path}")
    print(f"Sheets ({len(wb.sheetnames)}): {wb.sheetnames}")
    if path.suffix.lower() in (".xlsm", ".xltm"):
        print("NOTE: macro-enabled workbook; any edit must load with keep_vba=True or macros are destroyed.")

    for ws in wb.worksheets:
        print(f"\n== Sheet: {ws.title!r} ({ws.sheet_state})")
        print(f"   dimensions: {ws.dimensions} ({ws.max_row} rows x {ws.max_column} cols)")

        header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        if any(v is not None for v in header_cells):
            shown = [f"{get_column_letter(i + 1)}={v!r}" for i, v in enumerate(header_cells) if v is not None]
            print(f"   row 1: {', '.join(shown[:12])}{' ...' if len(shown) > 12 else ''}")

        formulas = sum(
            1
            for row in ws.iter_rows()
            for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("=")
        )
        print(f"   formula cells: {formulas}")
        if ws.merged_cells.ranges:
            print(f"   merged ranges: {[str(r) for r in ws.merged_cells.ranges]}")
        if getattr(ws, "tables", None):
            print(f"   tables: {list(ws.tables)}")
        # charts/images do not fully survive load+save in openpyxl; flag their presence
        charts = len(getattr(ws, "_charts", []) or [])
        images = len(getattr(ws, "_images", []) or [])
        if charts or images:
            print(f"   WARNING: {charts} chart(s), {images} image(s); openpyxl load+save can drop or alter these. Prefer edits that avoid resaving this sheet's file, or accept the loss explicitly.")
        if ws.freeze_panes:
            print(f"   freeze panes: {ws.freeze_panes}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
